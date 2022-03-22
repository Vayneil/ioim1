import math
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl import Workbook



def sigma_p(a, e, T, e_dot):
    R = 8.314
    W = math.exp(-a[6] * e)
    T1 = math.exp(a[3] / (R * (T + 273)))
    T2 = math.exp(a[5] / (R * (T + 273)))
    e1 = e ** a[1]
    edot1 = e_dot ** a[2]
    result = W * a[0] * T1
    result = result + ((1 - W) * a[4] * T2)
    result = result * edot1
    return result


def objective(a, e, T, e_dot, sigma_test):
    sum = 0
    for i in range(len(T)):
        for j in range(len(e[0])):
            err_squared = (sigma_test[i][j] - sigma_p(a, e[i][j], T[i], e_dot[i])) ** 2
            err_relative = err_squared / sigma_test[i][j]
            sum = sum + err_relative
    print(sum)
    return sum


def hooke_jeeves(x, s, alpha, epsilon, e, T, e_dot, sigma_test, constraints):
    while s > epsilon:
        xb = deepcopy(x)
        objective_values.append(objective(x, e, T, e_dot, sigma_test))
        x = trial(xb, s, constraints)
        if objective(x, e, T, e_dot, sigma_test) < objective(xb, e, T, e_dot, sigma_test):
            while True:
                xb1 = deepcopy(xb)
                xb = deepcopy(x)
                # temp = [2 * x for x in xb]
                x = [2 * x for x in xb]
                temp = []
                for i in range(len(x)):
                    temp.append(x[i] - xb1[i])
                    # x[i] = temp
                    if not constraints[1][i] > temp[i] > constraints[0][i]:
                        print('limip ipsum')
                        break
                else:
                    x = deepcopy(temp)
                # x = 2 * xb - xb1
                x = trial(x, s, constraints)
                if objective(x, e, T, e_dot, sigma_test) >= objective(xb, e, T, e_dot, sigma_test):
                    break
            x = deepcopy(xb)
        else:
            s = alpha * s
    return xb


def trial(x, s, constraints):
    for j in range(len(x)):
        trial_x = deepcopy(x)
        trial_x[j] = trial_x[j] + s * constraints[1][j]
        if objective(trial_x, e, T, e_dot, sigma_test) < objective(x, e, T, e_dot, sigma_test) and \
                constraints[1][j] > trial_x[j] > constraints[0][j]:
            x = deepcopy(trial_x)
        else:
            trial_x[j] = trial_x[j] - 2 * s * constraints[1][j]
            if objective(trial_x, e, T, e_dot, sigma_test) < objective(x, e, T, e_dot, sigma_test) and \
                    constraints[1][j] > trial_x[j] > constraints[0][j]:
                x = deepcopy(trial_x)
    print(x)
    return x


# load Excel file and initialize arrays
wb = load_workbook('./Doswiadczenia.xlsx', read_only=False)
e = []
sigma_test = []
T = []
e_dot = []

# set experimental values
for ws in wb.worksheets:
    e_values = [ws.cell(i, 1).value for i in range(3, 23)]
    sigma_values = [ws.cell(i, 2).value for i in range(3, 23)]
    T.append(ws.cell(2, 4).value)
    e_dot.append(ws.cell(2, 5).value)
    e.append(e_values)
    sigma_test.append(sigma_values)

a = [500, 0.5, 0.5, 5000, 0.5, 45000, 0.5]
constraints = [[1.0, 0.0, 0.0, 1.0, 0.0, 1.0, 0.0], [1000.0, 1.0, 1.0, 10000.0, 1.0, 90000.0, 1.0]]
# relative step size
s = 0.1
# 0 < alpha < 1
alpha = 0.75
# precision
epsilon = 0.0001
objective_values = []
result_a = hooke_jeeves(a, s, alpha, epsilon, e, T, e_dot, sigma_test, constraints)
result_objective = objective(result_a, e, T, e_dot, sigma_test)
# print(result_a)

# ws1 = wb.create_sheet(title="Results")
wb1 = load_workbook('./Results.xlsx', read_only=False)
ws1 = wb.worksheets[0]
ws1.cell(1, 3, result_objective)
for i in range(len(result_a)):
    ws1.cell(i + 1, 1, result_a[i])
for i in range(len(objective_values)):
    ws1.cell(i + 1, 2, objective_values[i])


wb1.save('./Results.xlsx')
wb.close()
wb1.close()

# 127.41171114187962
# [179.55596276914937, 0.5, 0.12200341075755708, -1014.2046975999997, 5.8713463056566155, 34944.01339230601, 0.23720028880244573]
# 127.41216971968127
# [179.56598455445413, 0.5, 0.12206022764828782, -1014.2046975999997, 5.8713463056566155, 34938.16057412525, 0.2373760505942261]
